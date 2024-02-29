import random
from statistics import mean
import signal
import time
import requests
import openpyxl


def handler(signum, frame): #hay que pararlo con CTRL+C o llegue a la solución fitness = 0, atrapa las interrupciones CTRL+C
    endTimes = time.time() - start_time
    print(endTimes)
    meanBook.save('../processed/myMeanBooklet.xlsx')
    minimumBook.save('../processed/myMinimumBooklet.xlsx')
    exit(1)


signal.signal(signal.SIGINT, handler)
start_time = time.time()
n = 100  # number of individuals per generation
m = 18  # number of tournament competitors
pepper = 2  # percentage of mutation
meanBook = openpyxl.load_workbook('../processed/myMeanBooklet.xlsx')
minimumBook = openpyxl.load_workbook('../processed/myMinimumBooklet.xlsx') #introduce la información en hojas de calculo de forma automática, una para medias, otra para minimos
meSheet = meanBook['pruebas']
miSheet = minimumBook['pruebas']
column = meSheet.max_column + 1 #nueva columna
c1 = meSheet.cell(row=1, column=column)
c2 = miSheet.cell(row=1, column=column)
c1.value = "n=" + str(n) + " m=" + str(m) + " p=" + str(pepper)
c2.value = "n=" + str(n) + " m=" + str(m) + " p=" + str(pepper) #introduce los datos de la prueba
row = 2

website = "UNDEFINED"

myChromosomes = []
myFitness = []


def firstGen():
    for c in range(0, n):
        myChromosomes.append(list('0000000000000000000000000000000000000000000000000000000000000000'))
        myFitness.append(0)
        cr = myChromosomes[c]
        for y in range(0, 64):
            if random.randrange(10) % 2 == 0:
                cr[y] = '0'
            else:
                cr[y] = '1'


def evaluation(chromosomes):
    for c in range(0, n):
        sndcr = "".join(chromosomes[c])
        req = requests.get(website + sndcr)
        myFitness[c] = float(req.text)
    meann = round(mean(myFitness), 2)
    print(meann)
    meann = str(meann)
    c1.value = meann #introduce el dato en excel


def selection():
    myNewGen = []
    for c in range(0, n):
        myList = []
        myTournament = []
        for a in range(0, m):
            tmp = random.randrange(n)
            myTournament.append(float(myFitness[tmp]))
            myList.append(tmp)
        myWinner = myFitness.index(min(myTournament))#toma el index del ganador en la lista de fitness para poder saber cual es el cromosoma ganador
        myNewGen.append(list(myChromosomes[myWinner]))

    return link(myNewGen)


def link(myNewGen): #metodo que se encarga de la reproduccion
    half = n / 2
    for a in range(0, int(n / 2)):
        cr1 = myNewGen[a]
        cr2 = myNewGen[a * 2]
        for p in range(0, 64):
            gen = random.randrange(10)
            if gen % 2 == 0:
                if cr1[p] == '0' and cr2[p] == '1':
                    cr1[p] = '1'
                    cr2[p] = '0'
                elif cr1[p] == '1' and cr2[p] == '0':
                    cr1[p] = '0'
                    cr2[p] = '1'
    return mutation(myNewGen)


def mutation(myNewGen):
    for c in myNewGen:
        for p in range(0, 64):
            gen = random.randrange(100)
            if 50 - (pepper / 2) < gen < 50 + (pepper / 2):
                if c[p] == '0':
                    c[p] = '1'
                else:
                    c[p] = '0'
    return myNewGen


firstGen()
minimum = 100
while minimum != 0.0:
    c1 = meSheet.cell(row=row, column=column)
    c2 = miSheet.cell(row=row, column=column) #cambia de celda en el excel
    evaluation(myChromosomes)
    minimum = min(myFitness)
    print(minimum)
    minium = str(minimum) + "\n"
    c2.value = minium
    row = row + 1
    myChromosomes = selection()
times = time.time() - start_time
print(times)
meanBook.save('../processed/myMeanBooklet.xlsx')
minimumBook.save('../processed/myMinimumBooklet.xlsx') #cuando acaba guarda los cambios en el documento
